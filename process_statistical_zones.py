#!/usr/bin/env python3
"""
Process CBS socioeconomic data per statistical zone and create mapping files.

Reads:
- T12 from CBS socio_eco21 publication 1955 (statistical areas within municipalities)
- T01 for local authority level data
- T07/T08 for regional council localities
- statistical_zones.json (from CBS 2011 GDB download)
- station_zone_mapping.json (from point-in-polygon matching)

Creates:
- site/data/statistical_zones_socioeconomic.json: per-zone socioeconomic data
- site/data/station_socioeconomic.json: per-station socioeconomic cluster
"""

import json
import openpyxl
from pathlib import Path

DATA_DIR = Path("/Users/harel/Developer/elections-vote-transfer/data")
SITE_DATA = Path("/Users/harel/Developer/elections-vote-transfer/site/data")


def parse_t12():
    """Parse T12: statistical areas within municipalities and local councils."""
    wb = openpyxl.load_workbook(DATA_DIR / "socio_eco21_T12.xlsx")
    ws = wb.active

    zones = {}
    for row in ws.iter_rows(min_row=9, max_col=7, values_only=True):
        settlement_code = row[0]
        settlement_name = row[1]
        stat_area = row[2]
        population = row[3]
        index_value = row[4]
        rank = row[5]
        cluster = row[6]

        if settlement_code is None or not isinstance(settlement_code, (int, float)):
            continue
        if stat_area is None or not isinstance(stat_area, (int, float)):
            continue

        settlement_code = int(settlement_code)
        stat_area = int(stat_area)
        yishuv_stat = settlement_code * 10000 + stat_area

        zones[str(yishuv_stat)] = {
            "cluster": int(cluster) if cluster is not None else None,
            "settlement": settlement_name.strip() if settlement_name else "",
            "settlement_code": settlement_code,
            "stat_area": stat_area,
            "index_value": round(index_value, 4) if index_value is not None else None,
            "rank": int(rank) if rank is not None else None,
            "population": round(population) if population is not None else None,
        }

    return zones


def parse_t01():
    """Parse T01: local authorities (settlement-level clusters)."""
    wb = openpyxl.load_workbook(DATA_DIR / "socio_eco21_T01.xlsx")
    ws = wb.active

    authorities = {}
    for row in ws.iter_rows(min_row=10, max_col=7, values_only=True):
        # A: municipal status, B: settlement code, C: name, D: population, E: index value, F: rank, G: cluster
        code = row[1]
        name = row[2]
        cluster = row[6]

        if code is None or not isinstance(code, (int, float)):
            # Some rows have name in C without code in B; also average row
            continue

        code = int(code)
        authorities[code] = {
            "cluster": int(cluster) if cluster is not None else None,
            "settlement": name.strip() if name else "",
        }

    return authorities


def parse_t07_t08():
    """Parse T07/T08: localities within regional councils."""
    localities = {}

    for t in ["T07", "T08"]:
        wb = openpyxl.load_workbook(DATA_DIR / f"socio_eco21_{t}.xlsx")
        ws = wb.active

        for row in ws.iter_rows(min_row=10, max_col=14, values_only=True):
            # F: settlement code, G: name, M: cluster 2019
            code = row[5]  # F
            name = row[6]  # G
            cluster = row[12]  # M

            if code is None or not isinstance(code, (int, float)):
                continue

            code = int(code)
            if code not in localities:
                localities[code] = {
                    "cluster": int(cluster) if cluster is not None else None,
                    "settlement": name.strip() if name else "",
                }

    return localities


def main():
    print("=== Parsing T12 (statistical areas within municipalities) ===")
    zones = parse_t12()
    print(f"  Found {len(zones)} statistical areas with socioeconomic data")

    # Get unique settlements from zones
    zone_settlements = set()
    for z in zones.values():
        zone_settlements.add(z["settlement_code"])
    print(f"  Covering {len(zone_settlements)} settlements (municipalities/local councils)")

    # Also parse settlement-level data for fallback
    print("\n=== Parsing T01 (local authorities) ===")
    authorities = parse_t01()
    print(f"  Found {len(authorities)} local authorities")

    print("\n=== Parsing T07/T08 (regional council localities) ===")
    regional = parse_t07_t08()
    print(f"  Found {len(regional)} localities within regional councils")

    # Load ArcGIS zones to enrich with socioeconomic data
    print("\n=== Loading ArcGIS statistical zones ===")
    with open(SITE_DATA / "statistical_zones.json", 'r', encoding='utf-8') as f:
        arcgis_data = json.load(f)
    arcgis_zones = arcgis_data["zones"]
    print(f"  Loaded {len(arcgis_zones)} zones from ArcGIS")

    # Match ArcGIS zones with CBS socioeconomic data
    matched = 0
    unmatched = 0
    enriched_zones = {}

    for yishuv_stat, zone in arcgis_zones.items():
        entry = {
            "cluster": None,
            "settlement": zone["settlement"],
            "settlement_code": zone["settlement_code"],
            "stat_area": zone["stat_area"],
        }

        # Try exact match in T12
        if yishuv_stat in zones:
            cbs = zones[yishuv_stat]
            entry["cluster"] = cbs["cluster"]
            entry["index_value"] = cbs["index_value"]
            entry["rank"] = cbs["rank"]
            entry["population"] = cbs["population"]
            matched += 1
        else:
            # Fallback to settlement-level cluster from T01 or T07/T08
            code = zone["settlement_code"]
            if code in authorities:
                entry["cluster"] = authorities[code]["cluster"]
                entry["source"] = "settlement"
            elif code in regional:
                entry["cluster"] = regional[code]["cluster"]
                entry["source"] = "regional"
            else:
                unmatched += 1

        enriched_zones[yishuv_stat] = entry

    print(f"\n  Matched stat area → CBS T12: {matched}")
    print(f"  Unmatched (no cluster): {unmatched}")
    fallback_settlement = sum(1 for z in enriched_zones.values() if z.get("source") == "settlement")
    fallback_regional = sum(1 for z in enriched_zones.values() if z.get("source") == "regional")
    print(f"  Fallback to settlement-level: {fallback_settlement}")
    print(f"  Fallback to regional locality: {fallback_regional}")

    # Save enriched zones
    output = {"zones": enriched_zones}
    out_path = SITE_DATA / "statistical_zones_socioeconomic.json"
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False)
    size_kb = out_path.stat().st_size / 1024
    print(f"\n  Saved: {out_path} ({size_kb:.0f} KB, {len(enriched_zones)} zones)")

    # Stats on cluster distribution
    cluster_counts = {}
    for z in enriched_zones.values():
        c = z.get("cluster")
        cluster_counts[c] = cluster_counts.get(c, 0) + 1
    print("\n  Cluster distribution:")
    for c in sorted(cluster_counts.keys(), key=lambda x: (x is None, x)):
        print(f"    Cluster {c}: {cluster_counts[c]} zones")

    # Also create a station → socioeconomic mapping combining station_zone_mapping + zones
    print("\n=== Creating station socioeconomic mapping ===")
    with open(SITE_DATA / "station_zone_mapping.json", 'r', encoding='utf-8') as f:
        station_zones = json.load(f)

    station_socio = {}
    matched_stations = 0
    for station_key, yishuv_stat in station_zones.items():
        ys = str(yishuv_stat)
        if ys in enriched_zones and enriched_zones[ys].get("cluster") is not None:
            station_socio[station_key] = {
                "zone": yishuv_stat,
                "cluster": enriched_zones[ys]["cluster"],
            }
            matched_stations += 1

    print(f"  Stations with socioeconomic cluster: {matched_stations} / {len(station_zones)}")

    # Save compact station socioeconomic data
    stn_path = SITE_DATA / "station_socioeconomic.json"
    with open(stn_path, 'w', encoding='utf-8') as f:
        json.dump(station_socio, f, ensure_ascii=False)
    size_kb = stn_path.stat().st_size / 1024
    print(f"  Saved: {stn_path} ({size_kb:.0f} KB)")

    print("\n=== Summary ===")
    print(f"  Statistical zones with cluster data: {sum(1 for z in enriched_zones.values() if z.get('cluster') is not None)} / {len(enriched_zones)}")
    print(f"  Stations matched to zones: {len(station_zones)} / 23551")
    print(f"  Stations with socioeconomic cluster: {matched_stations}")


if __name__ == "__main__":
    main()
